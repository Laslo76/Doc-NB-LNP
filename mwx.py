from openpyxl import load_workbook
import os


def fio(str_fio: str, order: str = 'fio') -> str:
    """
    Получить из полного ФИО фамилию и инициалы.
    Параметры:
        str_fio - строка полные фамилия и мя и отчество
        order - строка парядок выедения (fio - фамилия и иницыалы, iof - иницыалы фамилия)
    """
    list_fio = str_fio.split()
    if order == 'fio':
        return f'{list_fio[0]} {list_fio[1][0]}.{list_fio[2][0]}.'
    else:
        return f'{list_fio[1][0]}.{list_fio[2][0]}. {list_fio[0]}'


class MWX:
    def __init__(self) -> None:
        self.book = load_workbook('NBL.xlsx')

    def search_row(self, page_name: str = 'Docs', list_search: list = None) -> int:
        """
        Поиск строки удавлетворяющей критериям поиска list_search.
        Параметры:
            list_search - список списков [колонка, значение]
        """
        if list_search is None:
            list_search = list()
        result_search = 0
        sheet = self.book[page_name]
        for row in range(2, sheet.max_row + 1):
            flag = True
            for test_col in list_search:
                flag = flag and (str(sheet.cell(row, test_col[0]).value) == str(test_col[1]))
            if flag:
                result_search = row
                break
        return result_search

    def giv_const(self, name: str, default: str = None) -> str:
        """
        Возвращает значение константы. Константы собраны на странице Const.
        Параметры:
            name - содержит имя константы значение которой нужно вернуть (всегда первый столбец).
            default - возвращаемое значение в случае неудачного поиска. По умолчанию - None
        """
        name_sheet = 'Const'
        list_search = [[1, name]]
        crow = self.search_row(name_sheet, list_search)
        if crow == 0:
            return default
        else:
            sheet = self.book[name_sheet]
            return str(sheet.cell(crow, 2).value)

    def get_value(self, page_name: str = '', row: int = 0, col: int = 0, default: str = '') -> str:
        value = default
        if page_name == '' or row == 0 or col == 0:
            return value
        try:
            value = self.book[page_name].cell(row, col).value
        except:
            pass
        return value

    def get_name(self, page_name: str, list_search: list = None, default: str = '') -> str:
        if list_search is None:
            list_search = list()
        row = self.search_row(page_name, list_search)
        if 0 == row:
            return default
        else:
            sheet = self.book[page_name]
            return sheet.cell(row, 2).value

    def search_product(self, cod_: str) -> dict:
        """Осуществляет поиск товара по артикулу и возвращает, в случае успеха,
        справочник заполненный информацией о товаре. В противном  случае
        возвращаемый справочник будет пустым."""
        name_sheet = 'product'
        dict_ret = {'код': '', 'наименование': '', 'единица': '', 'ОКЕИ': ''}
        list_search = [[1, cod_]]
        row = self.search_row(name_sheet, list_search)
        if row != 0:
            sheet = self.book[name_sheet]
            keys = list(dict_ret.keys())
            for i in range(4):
                dict_ret[keys[i]] = sheet.cell(row, i + 1).value
        return dict_ret

    def search_nds(self, cod_: str) -> dict:
        """
        Осуществляет поиск ставки НДС по коду, в случае успеха, возвращает
        справочник заполненный информацией о ставке и порядке расчета. В противном
        случае возвращаемый справочник будет пустым.
        """
        name_sheet = 'NDS'
        dict_ret = {'наименование': '', 'представление': '', 'ставка': 0, 'ндс_сверху': 0}
        list_search = [[1, cod_]]
        row = self.search_row(name_sheet, list_search)
        if row != 0:
            sheet = self.book[name_sheet]
            keys = list(dict_ret.keys())
            for i in range(4):
                dict_ret[keys[i]] = sheet.cell(row, i + 1).value
        return dict_ret

    def total_customer(self, name_page: str = '', row: int = 0, ret_type: str = 'all') -> str:
        """
        Возвращает форматированную строку с информацией о контрагенте.
        Парметры:
            name_page
            row - номер строки
            ret_type - тип набора возвращаемых данных.
        """
        name_page_customer = 'counterparty'
        sheet = self.book[name_page_customer]

        list_search = [[1, self.get_value(name_page, row, 1)], [9, self.get_value(name_page, row, 22)]]
        crow = self.search_row(name_page_customer, list_search)

        if crow > 0:
            inn = sheet.cell(crow, 1).value
            name = sheet.cell(crow, 2).value
            address = sheet.cell(crow, 4).value
            phone = sheet.cell(crow, 3).value
            count_rs = sheet.cell(crow, 7).value
            count_ks = sheet.cell(crow, 8).value
            bik = sheet.cell(crow, 5).value
            bank = sheet.cell(crow, 6).value
            kpp = sheet.cell(crow, 9).value
            if ret_type == 'all':
                return f'{name}, ИНН {inn}, {address}, тел.: {phone}, р/с {count_rs},' \
                       f'в банке {bank}, БИК {bik}, к/с {count_ks}'
            elif ret_type == 'addr':
                return f'{address}'
            elif ret_type == 'trans':
                return f'{name}, ИНН {inn}'
            elif ret_type == 'name':
                return f'{name}'
            elif ret_type == 'key':
                return f'{inn}/{kpp}'
            elif ret_type == 'name_addr':
                return f'{name}, {address}'
        return f''

    def test_double(self, name_page: str = '', current_row: int = 0, list_test: list = None) -> int:
        """
        Производит подсчет строк идентичных заданной.
        Параметры:
            name_page - имя стираницы на которой производится подсчет
            current_row - номер строки, дубли которой считаем
            list_test - список стобцов по которым проверяем идентичность. Задается
                паврой - колонка, значение.
        """
        count_double = 0
        if name_page == '':
            return count_double
        sheet = self.book[name_page]

        if list_test is None:
            list_test = list()

        for row in range(2, sheet.max_row + 1):
            flag = True
            for test_col in list_test:
                flag = flag and (str(sheet.cell(row, test_col[0]).value) == str(test_col[1]))
            if flag:
                if current_row != row:
                    count_double += 1
        return count_double

    def transform(self, sheet_in: str, row: int, name_page: str = "", default: str = '', tmpl: str = '') -> str:
        """
        Функция возвращает преобразованное значение.
        """
        result = default
        if name_page == 'counterparty':
            sheet = self.book[sheet_in]
            list_search = [[1, sheet.cell(row, 3).value], [9, sheet.cell(row, 22).value]]
            result = self.get_name(name_page, list_search, default)

        return f'{result:{tmpl}}'

    def get_str_list(self, name_list='Docs', rows: list = None, columns: list = None) -> list:
        if columns is None:
            columns = list()
        s_page = self.book[name_list]
        if rows is None:
            rows = [x for x in range(2, s_page.max_row + 1)]
        list_docs = [tuple([self.transform(name_list, crow, col['trans'],
                                           self.get_value(name_list, crow, col['position']), col['format'])
                            for col in columns]) for crow in rows]
        return list_docs

    def save_list(self, name_sheet='', row: int = 0, list_data=None) -> None:
        if list_data is None:
            list_data = list()
        sheet = self.book[name_sheet]
        if row == 0:
            row = sheet.max_row + 1
        for column in range(len(list_data)):
            sheet.cell(row, column + 1).value = list_data[column]
        self.book.save('NBL.xlsx')

    def del_row(self, name_sheet, row):
        sheet = self.book[name_sheet]
        sheet.delete_rows(row)
        self.save_list(name_sheet)

    def prn_t12(self, row_print: int) -> None:
        """
        Вывод документа по форме T12
        Параметры:
            row_print - номер строки документа из закладки Docs
        """
        page = self.book['Docs']

        template_sf = load_workbook('торг12.xlsx')
        target_page = template_sf.active

        customer = self.total_customer(str(row_print))
        dic_str = self.search_product(str(page.cell(row_print, 14).value))
        dic_nds = self.search_nds(str(page.cell(row_print, 13).value))

        target_page.cell(8, 4).value = customer
        target_page.cell(12, 4).value = customer
        target_page.cell(13, 42).value = f'{page.cell(row_print, 17).value}'
        target_page.cell(15, 42).value = f'{page.cell(row_print, 18).value}'
        target_page.cell(14, 4).value = f'{page.cell(row_print, 17).value} от {page.cell(row_print, 18).value}'
        target_page.cell(17, 12).value = page.cell(row_print, 1).value
        target_page.cell(17, 16).value = page.cell(row_print, 4).value
        target_page.cell(17, 16).number_format = 'DD.MM.YYYY'

        target_page.cell(23, 3).value = dic_str['наименование']
        target_page.cell(23, 8).value = dic_str['код']
        target_page.cell(23, 9).value = dic_str['единица']
        target_page.cell(23, 13).value = dic_str['ОКЕИ']

        quantity = page.cell(row_print, 15).value
        cost = page.cell(row_print, 16).value

        target_page.cell(23, 23).value = quantity
        target_page.cell(23, 26).value = cost

        target_page.cell(23, 34).value = dic_nds['представление']
        if dic_nds['ндс_сверху'] == 1:
            sum_bez_nds = float(quantity) * float(cost)
            nds = dic_nds['ставка'] / 100
            target_page.cell(23, 29).value = sum_bez_nds
            target_page.cell(23, 38).value = sum_bez_nds * nds
            target_page.cell(23, 41).value = sum_bez_nds * (1 + nds)
        else:
            sum_s_nds = quantity * cost
            nds = dic_nds['ставка'] / 100
            target_page.cell(23, 29).value = sum_s_nds / (1 + nds)
            target_page.cell(23, 38).value = sum_s_nds * nds / (1 + nds)
            target_page.cell(23, 41).value = sum_s_nds

        target_page.cell(45, 8).value = page.cell(row_print, 4).value

        path = f'{os.getcwd()}\\docs\\'
        template_sf.save(f'{path}{str(page.cell(row_print, 1).value)}_t12.xlsx')

    def prn_p4(self, row_print: int) -> None:
        """
        Вывод документа по форме P4
        :param row_print: число - номер строки из листа Docs
        :return: None
        """
        s_page = self.book['Docs']

        template_sf = load_workbook('Приложение4.xlsx')
        target_page = template_sf.active

        doc_date = s_page.cell(row_print, 4).value
        doc_number = s_page.cell(row_print, 1).value
        doc_prefix = self.giv_const('Префикс')
        doc_org = self.giv_const('НаименованиеПолное')
        doc_org_inn = self.giv_const('ИНН')
        doc_adr_pog = self.giv_const('Адрес отгрузки')
        doc_shipped = self.giv_const('Отпустил ФИО')
        doc_post_shipped = self.giv_const('Отпустил должность')
        doc_adr_raz = self.total_customer(s_page.cell(row_print, 3).value, row_print, 'addr')

        if not (doc_prefix is None):
            doc_number_prn = doc_prefix
        else:
            doc_number_prn = ""
        doc_number_prn += f'{doc_number}'

        customer = self.total_customer(s_page.cell(row_print, 3).value)
        dic_str = self.search_product(str(s_page.cell(row_print, 14).value))
        dic_nds = self.search_nds(str(s_page.cell(row_print, 13).value))

        target_page.cell(8, 4).value = doc_date
        target_page.cell(8, 32).value = doc_date
        target_page.cell(8, 18).value = doc_number_prn
        target_page.cell(8, 46).value = doc_number_prn
        # РАЗДЕЛ 2
        target_page.cell(17, 2).value = customer
        target_page.cell(19, 2).value = self.total_customer(s_page.cell(row_print, 3).value, row_print, 'addr')
        # РАЗДЕЛ 3
        target_page.cell(22, 2).value = dic_str['наименование']
        vr = float(s_page.cell(row_print, 19).value)
        vt = float(s_page.cell(row_print, 20).value)
        if dic_str['ОКЕИ'] == '112':
            vv = float(s_page.cell(row_print, 15).value)
            vm = vv * vr
        else:
            vm = float(s_page.cell(row_print, 15).value)
            vv = vm / vr

        target_page.cell(24, 2).value = f'Масса нефтепродукта-{vm:.3f}кг; объём нефтепродукта-{vv:.3f}дм3;' \
                                        f' плотность-{vr:.3f}кг/м3; температура {vt}°C'
        summa = float(s_page.cell(row_print, 16).value) * float(s_page.cell(row_print, 15).value)
        nds = dic_nds['ставка'] / 100
        if dic_nds['ндс_сверху'] == 1:
            target_page.cell(1, 1).value = summa * (1 + nds)
        else:
            target_page.cell(1, 1).value = summa
        # РАЗДЕЛ 4
        pass_cash = s_page.cell(row_print, 7).value
        const_inn = self.giv_const('ИНН')
        target_page.cell(31, 2).value = f'Паспорт качества: {pass_cash}, ИНН {const_inn}'
        target_page.cell(33, 2).value = f'№ {doc_number_prn} от {doc_date}; ИНН {const_inn}'
        # РАЗДЕЛ 5
        target_page.cell(36, 2).value = f'дата доставки: {doc_date}'
        target_page.cell(38, 30).value = f'Пломбы: {s_page.cell(row_print, 8).value}'
        # РАЗДЕЛ 6
        target_page.cell(41, 2).value = self.total_customer(s_page.cell(row_print, 21).value, row_print, 'trans')
        target_page.cell(41, 30).value = f'{s_page.cell(row_print, 9).value}'
        # РАЗДЕЛ 7
        target_page.cell(44, 2).value = f'{s_page.cell(row_print, 11).value}'
        target_page.cell(44, 30).value = f'{s_page.cell(row_print, 10).value}'
        # РАЗДЕЛ 8
        target_page.cell(54, 2).value = f'{doc_org}, ИНН {doc_org_inn}'
        target_page.cell(56, 2).value = f'{doc_org}, ИНН {doc_org_inn}'
        target_page.cell(58, 2).value = f'{doc_adr_pog}'
        target_page.cell(58, 28).value = f'{doc_date} {s_page.cell(row_print, 5).value}'
        target_page.cell(60, 2).value = f'{doc_date}'
        target_page.cell(60, 28).value = f'{doc_date}'
        target_page.cell(64, 2).value = s_page.cell(row_print, 12).value
        target_page.cell(68, 2).value = doc_post_shipped
        target_page.cell(68, 25).value = doc_shipped
        target_page.cell(68, 28).value = f'{s_page.cell(row_print, 9).value}'
        # РАЗДЕЛ 10
        target_page.cell(76, 2).value = doc_adr_raz
        target_page.cell(78, 2).value = f'{doc_date}'
        target_page.cell(78, 28).value = f'{doc_date}'
        target_page.cell(80, 28).value = s_page.cell(row_print, 12).value
        target_page.cell(82, 2).value = f'{vm:.3f}кг, {vv:.3f}л'
        target_page.cell(84, 28).value = f'{s_page.cell(row_print, 9).value}'

        path = f'{os.getcwd()}\\docs\\'
        template_sf.save(f'{path}{str(s_page.cell(row_print, 1).value)}_p4.xlsx')

    def prn_sf(self, row_print: int) -> None:
        """
        Выводит документ СЧЕТ-ФАКТУРУ
        :param row_print: число номер строки выводимого документа из листа Docs
        :return: None
        """
        s_page = self.book['Docs']

        dic_str = self.search_product(str(s_page.cell(row_print, 14).value))
        dic_nds = self.search_nds(str(s_page.cell(row_print, 13).value))

        template_sf = load_workbook('СФ.xlsx')
        target_page = template_sf.active

        target_page.cell(2, 5).value = s_page.cell(row_print, 2).value
        target_page.cell(2, 19).value = s_page.cell(row_print, 4).value

        target_page.cell(4, 5).value = self.giv_const('Наименование')
        target_page.cell(4, 35).value = self.total_customer(s_page.cell(row_print, 3).value, row_print, 'name')

        target_page.cell(5, 5).value = self.giv_const('Адрес')
        target_page.cell(5, 35).value = self.total_customer(s_page.cell(row_print, 3).value, row_print, 'addr')

        target_page.cell(6, 5).value = f'{self.giv_const("ИНН")}/{self.giv_const("КПП")}'
        target_page.cell(6, 35).value = self.total_customer(s_page.cell(row_print, 3).value, row_print, 'key')

        target_page.cell(8, 5).value = self.total_customer(s_page.cell(row_print, 3).value, row_print, 'addr')

        target_page.cell(9, 5).value = s_page.cell(row_print, 6).value
        target_page.cell(10, 5).value = f'№ п/п 1 №{s_page.cell(row_print, 1).value} от ' \
                                        f'{s_page.cell(row_print, 4).value} г.'

        quantity = s_page.cell(row_print, 15).value
        cost = s_page.cell(row_print, 16).value

        target_page.cell(15, 3).value = dic_str['наименование']
        target_page.cell(15, 14).value = dic_str['единица']
        target_page.cell(15, 13).value = dic_str['ОКЕИ']
        target_page.cell(15, 17).value = quantity
        target_page.cell(15, 20).value = cost
        target_page.cell(15, 30).value = dic_nds['представление']
        if dic_nds['ндс_сверху'] == 1:
            sum_bez_nds = float(quantity) * float(cost)
            nds = dic_nds['ставка'] / 100
            target_page.cell(15, 23).value = sum_bez_nds
            target_page.cell(15, 33).value = sum_bez_nds * nds
            target_page.cell(15, 36).value = sum_bez_nds * (1 + nds)
        else:
            sum_s_nds = quantity * cost
            nds = dic_nds['ставка'] / 100
            target_page.cell(15, 23).value = sum_s_nds / (1 + nds)
            target_page.cell(15, 33).value = sum_s_nds * nds / (1 + nds)
            target_page.cell(15, 36).value = sum_s_nds

        target_page.cell(18, 11).value = fio(self.giv_const('Отпустил ФИО'))
        target_page.cell(18, 32).value = fio(self.giv_const('Бухгалтер ФИО'))

        path = f'{os.getcwd()}\\docs\\'
        template_sf.save(f'{path}{str(s_page.cell(row_print, 1).value)}_sf.xlsx')


if __name__ == '__main__':
    row_doc_print = 2
    wb = MWX()
    print(wb.get_str_list(
        name_list='Docs',
        columns=[{'position': 1, 'format': '', 'trans': '', 'width': 50},
                 {'position': 4, 'format': '', 'trans': '', 'width': 75},
                 {'position': 3, 'format': '', 'trans': 'counterparty', 'width': 300}]))
